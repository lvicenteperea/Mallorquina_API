from datetime import datetime, date
import json

# Para trabajar con Excel PANDA
import pandas as pd

# Para trabajar con Excel OPENPYXL
from openpyxl import Workbook

from app.models.mll_cfg import obtener_cfg_general
from app.config.db_mallorquina import get_db_connection_mysql, close_connection_mysql
from app.services.auxiliares.sendgrid_service import enviar_email

from app.utils.utilidades import graba_log, imprime
from app.utils.InfoTransaccion import InfoTransaccion
from app.config.settings import settings
from app.utils.mis_excepciones import MiException

RUTA = f"{settings.RUTA_DATOS}cierre_caja/"

#----------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------
def proceso(param: InfoTransaccion) -> list:
    param.debug="Inicio"
    resultado = []
    datos = []
    config = obtener_cfg_general(param)
    # fecha = param.parametros[0] # el primer  atributo de InfArqueoCajaRequest
    entidad = param.parametros[1]  # el segundo atributo de InfArqueoCajaRequest

    param.debug = "get_db_connection_mysql"
    try:
        conn_mysql = get_db_connection_mysql()
        cursor_mysql = conn_mysql.cursor(dictionary=True)

        param.debug = "Select"
        if entidad == 0:
            query = "SELECT * FROM mll_cfg_entidades WHERE activo='S'"
            cursor_mysql.execute(query)
        else:
            query = "SELECT * FROM mll_cfg_entidades WHERE id = %s AND activo='S'"
            cursor_mysql.execute(query, (entidad,))

        lista_entidades = cursor_mysql.fetchall()

        for entidad in lista_entidades:
            imprime([f"Procesando TIENDA: {entidad['ID']}-{entidad['Nombre']}. De la BBDD: {entidad['id_bbdd']}"], "-")
            datos.append(consultar(param, entidad["ID"], conn_mysql))

        if datos:
            resultado.append(a_excel_con_pd(param, datos))

            if param.ret_code == 0:
                resultado.append(a_excel_con_openpyxl(param, datos))

            return resultado
        else:
            return ["No se ha generado fichero porque no hay datos"]



    except MiException as e:
        param.error_sistema(e=e, debug="arqueo_caja_info.Proceso.MiExcepcion")
        raise e
    except Exception as e:
        param.error_sistema(e=e, debug="arqueo_caja_info.Proceso.Excepcion")
        raise e # HTTPException(status_code=500, detail=e)


    finally:
        close_connection_mysql(conn_mysql, cursor_mysql)

        enviar_email(config["Lista_emails"],
                     "Proceso finalizado",
                     ["El proceso de informes de arqueo de caja ha terminado."]
        )


#----------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------
def consultar(param: InfoTransaccion, id_entidad, conn_mysql) -> list:
    resultado = []
    param.debug="Inicio"
    fecha = param.parametros[0]

    param.debug = "get_db_connection_mysql"
    try:
        cursor_mysql = conn_mysql.cursor(dictionary=True)

        param.debug = "Select"
        query = f"""SELECT 
                        vd.id_entidad,
                        t.nombre Tienda,
                        vd.serie,
                        pf.descripcion Nombre_TPV,
                        vd.fecha,
                        vd.cierre_tpv_id,
                        vd.cierre_tpv_desc,
                        vmp.id_medios_pago,
                        mp.nombre Nombre_MdP,
                        sum(vd.imp_arqueo_ciego) AS total_arqueo_ciego,
                        SUM(vmp.ventas) AS total_ventas,
                        SUM(vmp.operaciones) AS total_operaciones
                    FROM mll_rec_ventas_diarias vd
                        JOIN  mll_rec_ventas_medio_pago vmp ON vd.id = vmp.id_ventas_diarias
                    LEFT JOIN mll_cfg_entidades t         ON vd.id_entidad = t.id
                    LEFT JOIN tpv_puestos_facturacion pf ON vd.serie = pf.serie and vd.id_entidad = pf.id_entidad
                    LEFT JOIN mll_mae_medios_pago mp ON vmp.id_medios_pago = mp.id
                    where vd.id_entidad={id_entidad}
                      and vd.fecha = STR_TO_DATE('{fecha}', '%Y-%m-%d')
                    GROUP BY 
                        vd.id_entidad,
                        t.nombre,
                        vd.serie,
                        pf.descripcion,
                        vd.fecha,
                        vd.cierre_tpv_id,
                        vd.cierre_tpv_desc,
                        vmp.id_medios_pago,
                        mp.nombre
                 """

        param.debug="execute del cursor"
        cursor_mysql.execute(query)
        datos = cursor_mysql.fetchall()

        param.debug= "en el FOR"
        for row in datos:
            resultado.append(row)

        return resultado

    except Exception as e:
        param.error_sistema(e=e, debug="proceso.Exception")
        raise 




#----------------------------------------------------------------------------------------
# Creamos el escritor de Excel con la librería PANDA
#----------------------------------------------------------------------------------------
def a_excel_con_pd(param: InfoTransaccion, todos_los_conjuntos):
    param.debug = "Inicio"
    nombre_fich = "resultado_panda.xlsx"
    fichero = f"{RUTA}{nombre_fich}"

    try:
        with pd.ExcelWriter(fichero) as writer:
            param.debug = "Bucle for"
            for sublista in todos_los_conjuntos:
                # Si la sublista está vacía, pasamos de largo
                if not sublista:
                    continue
                
                param.debug = "Bucle for 1"
                # 1. Convertimos la sublista (lista de dicts) a DataFrame
                df = pd.DataFrame(sublista)

                param.debug = "Bucle for 2"
                # 2. Eliminamos las columnas que NO queremos exportar
                columnas_a_eliminar = ["id_entidad", "id_tpv", "id_medios_pago"]
                df.drop(columns=columnas_a_eliminar, axis=1, errors='ignore', inplace=True)

                param.debug = "convertimos fecha"
                # 3. Convertimos 'fecha' a formato dd/mm/aaaa
                #    Primero la parseamos a datetime y luego la formateamos
                df["fecha"] = pd.to_datetime(df["fecha"]).dt.strftime('%d/%m/%Y')

                # 4. Convertimos 'total_ventas' y 'total_operaciones' a floats, luego a strings con coma decimal
                df["total_ventas"] = df["total_ventas"].astype(float).map(
                    lambda x: f"{x:.2f}".replace('.', ',')  # 2 decimales con coma
                )
                df["total_operaciones"] = df["total_operaciones"].astype(float).map(
                    lambda x: f"{x:.0f}"  # sin decimales (asumiendo que es un entero), 
                                        # si quisieras 2 decimales: f"{x:.2f}".replace('.', ',')
                )

                param.debug = "Bucle for Nombre"
                # 5. Nombramos la hoja según la columna "Tienda" (limitamos a 31 caracteres para Excel)
                nombre_tienda = df["Tienda"].iloc[0]
                nombre_hoja = nombre_tienda[:31]

                # 6. Exportamos este DataFrame a una hoja en el Excel
                df.to_excel(writer, sheet_name=nombre_hoja, index=False)

        return {"fichero": f'{nombre_fich}', "texto": f"'{nombre_fich}' generado correctamente"}
    
    except Exception as e:
        param.error_sistema(e=e, debug="proceso.Exception")
        raise  e

#----------------------------------------------------------------------------------------
# Creamos el escritor de Excel con la librería PANDA
#----------------------------------------------------------------------------------------
def a_excel_con_openpyxl(param: InfoTransaccion, todos_los_conjuntos):
    param.debug ="Inicio"
    try:
        # 1. Creamos el libro de Excel y removemos la hoja por defecto
        wb = Workbook()
        ws_default = wb.active
        wb.remove(ws_default)
        nombre_fich = "resultado_openpyxl.xlsx"
        fichero = f"{RUTA}{nombre_fich}"
        

        param.debug = "2. Elimiar Columnas"
        # 2. Columnas que NO queremos mostrar
        columnas_excluir = {"id_entidad", "id_tpv", "id_medios_pago"}

        for sublista in todos_los_conjuntos:
            # Si la sublista está vacía, la saltamos
            if not sublista:
                continue
            
            param.debug = "3. Procesamos"
            # 3. Preprocesamos la sublista para convertir valores y eliminar columnas
            #    - Convertimos "fecha" a datetime
            #    - Convertimos "total_ventas" y "total_operaciones" a float
            #    - Quitamos las columnas que no queremos
            datos_procesados = []
            # imprime([sublista], "*")
            for fila in sublista:
                # Creamos una copia limpia
                nueva_fila = {}
                for k, v in fila.items():
                    if k in columnas_excluir:
                        continue  # saltar columnas no deseadas

                    if k == "fecha":
                        # Convertir "AAAA-MM-DD" a datetime
                        if isinstance(v, str):     # Si es una cadena, puedes usar strptime
                            nueva_fila[k] = datetime.strptime(v, "%Y-%m-%d").date()
                        elif isinstance(v, date): # Si ya es un objeto date, úsalo directamente
                            nueva_fila[k] = v
                        else:
                            raise ValueError(f"El valor {type(v)}-{v} no es una cadena ni un objeto date")
                    elif k in ("total_ventas", "total_operaciones"):
                        # Convertir a float
                        nueva_fila[k] = float(v)
                    else:
                        # Dejar el valor tal cual
                        nueva_fila[k] = v
                datos_procesados.append(nueva_fila)

            param.debug = "4. Obtenemos nombre tienda"
            # 4. Obtenemos el nombre de la tienda (primer elemento)
            nombre_tienda = datos_procesados[0]["Tienda"]
            # Máximo 31 caracteres para el nombre de la hoja
            sheet_name = nombre_tienda[:31]

            param.debug = "5. Creamos nueva hoja"
            # imprime([param.debug, sheet_name], "=")
            # 5. Creamos una hoja nueva con el nombre de la tienda
            ws = wb.create_sheet(title=sheet_name)
            
            param.debug = "6. Escribimos cabecera"
            # 6. Escribimos la CABECERA
            #    Obtenemos las columnas del primer registro (ordenadas)
            columnas = list(datos_procesados[0].keys())
            ws.append(columnas)  # primera fila con los nombres de columna

            param.debug = "7. Filas"
            # 7. Escribimos los datos fila por fila
            for fila in datos_procesados:
                row_data = [fila[col] for col in columnas]
                ws.append(row_data)

            param.debug = "8. Formatos"
            # 8. Aplicamos FORMATO a las celdas
            #    - Fecha en formato dd/mm/yyyy
            #    - total_ventas con 2 decimales
            #    - total_operaciones con 0 decimales (o lo que prefieras)
            
            # Calculamos los índices de columna para "fecha", "total_ventas", etc.
            # ojo: openpyxl usa 1-based index, ws["A1"] es row=1,col=1
            idx_fecha = columnas.index("fecha") + 1 if "fecha" in columnas else None
            idx_ventas = columnas.index("total_ventas") + 1 if "total_ventas" in columnas else None
            idx_operaciones = columnas.index("total_operaciones") + 1 if "total_operaciones" in columnas else None

            # Recorremos las filas de datos (comienzan en la 2, ya que la 1 es cabecera)
            for row in range(2, 2 + len(datos_procesados)):
                # Fecha (dd/mm/yyyy)
                if idx_fecha:
                    cell_fecha = ws.cell(row=row, column=idx_fecha)
                    # Indicamos que es un formato de fecha dd/mm/yyyy
                    cell_fecha.number_format = "DD/MM/YYYY"

                # total_ventas con 2 decimales
                if idx_ventas:
                    cell_ventas = ws.cell(row=row, column=idx_ventas)
                    cell_ventas.number_format = "#,##0.00"

                # total_operaciones sin decimales
                if idx_operaciones:
                    cell_oper = ws.cell(row=row, column=idx_operaciones)
                    cell_oper.number_format = "#,##0"

        param.debug = "9. Guardamos"
        # 9. Guardamos el archivo
        wb.save(fichero)

        return {"fichero": f'{nombre_fich}', "texto": f"'{nombre_fich}' generado correctamente"}

    except Exception as e:
        param.error_sistema(e=e, debug="proceso.Exception")
        raise 