import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

from app.utils.functions import graba_log
from fastapi import HTTPException
from app.utils.mis_excepciones import MadreException
from app.utils.InfoTransaccion import InfoTransaccion
from app import mi_libreria as mi

#----------------------------------------------------------------------------------------
#----------------------------------------------------------------------------------------
def proceso(param: InfoTransaccion) -> list:
    """
    Convierte un archivo Excel de origen al formato deseado y lo guarda en el destino.

    Args:
        origen_path (str): Ruta del archivo de origen.
        output_path (str): Ruta donde se guardará el archivo convertido.

        origen_path = "app/datos/export_sqlpyme.xlsx"
        destino_path = "app/datos/importa_TPV.xlsx"
    """
    resultado = []
    donde = "Inicio"

    try:
        path = "app/datos/tarifas_a_TPV/"
        
        if param.parametros and param.parametros[0]:
            origen_path = f"{path}{param.parametros[0]}"
        else:
            param.ret_txt = "No ha llegado fichero origen para crear el nuevo fichero"
            raise MadreException(param.to_dict())
                
        if len(param.parametros) >= 2  and param.parametros[1]:
            output_path = f"{path}{param.parametros[1]}"
        else:
            output_path = f"{path}tarifas_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        donde = "convierte_con_pd"
        resultado = convierte_con_pd(param, origen_path, output_path)
        
        # donde = "convierte_con_openpyxl" 
        #resulta do.append(convierte_con_openpyxl(param, f"{origen_path}", f"{path}tarifas_{datetime.now().strftime('%Y%m%d%H%M%S')}"))


    except MadreException as e:
        param.ret_code = -1

              
    except Exception as e:
        param.ret_code = -99
        param.ret_txt = "Error general. contacte con su administrador"
        graba_log({"ret_code": param.ret_code, "ret_txt": param.ret_txt}, f"Excepción tarifas_a_TPV.proceso-{donde}", e)

    
    finally:
        return resultado

# -----------------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------------
def convierte_con_pd (param, origen_path, output_path):
    resultado = []

    try:
        donde = "#Leer el archivo de origen"
        origen_df = pd.read_excel(origen_path)

        donde = "# Crear un DataFrame vacío con las columnas deseadas"
        columns_destino = [
            "Id Plato", "Descripcion", "Barra", "Comedor", "Terraza", "Hotel", 
            "Reservado", "Menú", "Orden Factura", "Orden Cocina", "OrdenTactil",
            "Grupo Carta 1", "Grupo Carta 2", "Grupo Carta 3", "Grupo Carta 4",
            "Familia", "Código Barras", "Centro", "Centro 2", "Centro 3"
        ]
        converted_df = pd.DataFrame(columns=columns_destino)

        donde = "# Mapear las columnas según las indicaciones"
        converted_df["Id Plato"] = origen_df["Código"]
        converted_df["Descripcion"] = origen_df["Nombre"]
        converted_df["Barra"] = origen_df["PVP TIENDA"]
        converted_df["Comedor"] = origen_df["PVP SALON TIENDAS"]
        converted_df["Terraza"] = origen_df["PVP TERRAZA QUEVEDO"]
        converted_df["Hotel"] = ""
        converted_df["Reservado"] = ""
        converted_df["Menú"] = ""
        converted_df["Orden Factura"] = ""
        converted_df["Orden Cocina"] = ""
        converted_df["OrdenTactil"] = ""
        converted_df["Grupo Carta 1"] = origen_df["GRUPO DE CARTA"]
        converted_df["Grupo Carta 2"] = ""
        converted_df["Grupo Carta 3"] = ""
        converted_df["Grupo Carta 4"] = ""
        converted_df["Familia"] = ""
        converted_df["Código Barras"] = origen_df["Código de barras"]

        donde = "# Mapear centros de preparación a 'X'"
        converted_df["Centro"] = origen_df["CENTRO PREPARACION CAFETERA"].apply(lambda x: "X" if not pd.isna(x) else "")
        converted_df["Centro 2"] = origen_df["CENTRO PREPARACION PLANCHA"].apply(lambda x: "X" if not pd.isna(x) else "")
        converted_df["Centro 3"] = origen_df["CENTRO PREPARACION COCINA"].apply(lambda x: "X" if not pd.isna(x) else "")

        # Guardar el DataFrame convertido en el archivo de destino"
        donde = f"Guarda el archivo {output_path}.   "
        converted_df.to_excel(output_path, index=False)
        
        donde = f"# Contar registros generados"
        num_registros_destino = converted_df.shape[0]
        num_registros_origen = origen_df.shape[0]
        
        if num_registros_origen == num_registros_destino:
            param.ret_code = 0
            param.ret_txt = f"Ok: Ambos archivos ({origen_path} y {output_path}) tienen {num_registros_origen} registros"
        else:
            param.ret_code = -1
            param.ret_txt = f"Error: El archivo origen ({origen_path}) tiene {num_registros_origen} registros. El archivo generado ({output_path}) tiene {num_registros_destino} registros."
        
        resultado = [param.ret_txt]

    except Exception as e:
        param.ret_code = -1
        param.ret_txt = "Error general. contacte con su administrador"
        graba_log({"ret_code": param.ret_code, "ret_txt": param.ret_txt}, f"Excepción tarifas_a_TPV.proceso-{donde}", e)
       

    finally:
        return resultado

'''
# -----------------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------------
def convierte_con_openpyxl (param, origen_path, output_path):
    resultado = []

    try:
        # Cargar el archivo de origen con openpyxl
        wb_origen = load_workbook(origen_path)
        ws_origen = wb_origen.active  # Suponemos que los datos están en la primera hoja

        # Crear un nuevo archivo para el destino
        wb_destino = Workbook()
        ws_destino = wb_destino.active
        ws_destino.title = "Datos Convertidos"

        # Definir las columnas del archivo destino
        columns_destino = [
            "Id Plato", "Descripcion", "Barra", "Comedor", "Terraza", "Hotel", 
            "Reservado", "Menú", "Orden Factura", "Orden Cocina", "OrdenTactil",
            "Grupo Carta 1", "Grupo Carta 2", "Grupo Carta 3", "Grupo Carta 4",
            "Familia", "Código Barras", "Centro", "Centro 2", "Centro 3"
        ]
        ws_destino.append(columns_destino)  # Escribir los encabezados

        # Mapeo de columnas de origen a destino
        for row in ws_origen.iter_rows(min_row=2, values_only=True):  # Saltar la fila de encabezados
            id_plato = row[0]  # "Código"
            descripcion = row[2]  # "Nombre"
            barra = row[5] if row[5] not in (None, "", 0) else ""  # "PVP TIENDA"
            comedor = row[6] if row[6] not in (None, "", 0) else ""  # "PVP SALON TIENDAS"
            terraza = row[7] if row[7] not in (None, "", 0) else ""  # "PVP TERRAZA QUEVEDO"
            hotel = ""
            reservado = ""
            menu = ""
            orden_factura = ""
            orden_cocina = ""
            orden_tactil = ""
            grupo_carta_1 = row[14]  # "GRUPO DE CARTA"
            grupo_carta_2 = ""
            grupo_carta_3 = ""
            grupo_carta_4 = ""
            familia = ""
            codigo_barras = row[1]  # "Código de barras"
            centro = "X" if row[15] else ""  # "CENTRO PREPARACION CAFETERA"
            centro_2 = "X" if row[16] else ""  # "CENTRO PREPARACION PLANCHA"
            centro_3 = "X" if row[17] else ""  # "CENTRO PREPARACION COCINA"

            # Crear la fila destino
            fila_destino = [
                id_plato, descripcion, barra, comedor, terraza, hotel,
                reservado, menu, orden_factura, orden_cocina, orden_tactil,
                grupo_carta_1, grupo_carta_2, grupo_carta_3, grupo_carta_4,
                familia, codigo_barras, centro, centro_2, centro_3
            ]
            ws_destino.append(fila_destino)

        # Ajustar alineación de celdas
        for row in ws_destino.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Guardar el archivo destino
        wb_destino.save(output_path)
        print(f"Archivo final generado con formato guardado en: {output_path}")

        # Contar registros
        num_registros_origen = ws_origen.max_row - 1
        num_registros_destino = ws_destino.max_row - 1
        print(f"El archivo de origen tiene {num_registros_origen} registros.")
        print(f"El archivo final generado tiene {num_registros_destino} registros.")


        donde = "Comparando registros"       
        if num_registros_origen == num_registros_destino:
            param.ret_code = 0
            param.ret_txt = f"Ok: Ambos archivos tienen {num_registros_origen} registros"
        else:
            param.ret_code = -1
            param.ret_txt = f"Error: El archivo origen tiene {num_registros_origen} registros. El archivo generado tiene {num_registros_destino} registros."
        
        resultado = [param.ret_txt]

    except Exception as e:
        param.ret_code = -1
        param.ret_txt = "Error general. contacte con su administrador"
        graba_log({"ret_code": param.ret_code, "ret_txt": param.ret_txt}, f"Excepción tarifas_a_TPV.proceso-{donde}", e)
       

    finally:
        return resultado
'''

